"""
Export & Reporting System for Freetrade Portfolio Dashboard

This module provides comprehensive export and reporting functionality:
- CSV export with auto-generated filenames
- Multi-sheet Excel export with formatting
- Professional HTML reports with embedded charts
- PDF report generation (requires weasyprint)
- Export panel UI for Streamlit
"""

import streamlit as st
import pandas as pd
import numpy as np
import io
from datetime import datetime

# Optional Excel export
try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
    from openpyxl.utils.dataframe import dataframe_to_rows
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

# Optional PDF export
try:
    from weasyprint import HTML as WeasyHTML
    WEASYPRINT_AVAILABLE = True
except ImportError:
    WEASYPRINT_AVAILABLE = False

# For charts in HTML reports
try:
    import plotly.express as px
    PLOTLY_AVAILABLE = True
except ImportError:
    PLOTLY_AVAILABLE = False


def export_portfolio_csv(holdings_df, account_name="Portfolio", filename=None):
    """
    Export holdings to CSV format.

    Args:
        holdings_df: DataFrame with holdings data
        account_name: Name of the account (ISA/SIPP/Combined)
        filename: Optional custom filename

    Returns:
        Tuple of (csv_data, filename) for download
    """
    if holdings_df is None or holdings_df.empty:
        return None, None

    # Auto-generate filename with date if not provided
    if filename is None:
        date_str = datetime.now().strftime("%Y-%m-%d")
        filename = f"portfolio_{account_name}_{date_str}.csv"

    # Prepare data for export
    export_df = holdings_df.copy()

    # Convert to CSV
    csv_buffer = io.StringIO()
    export_df.to_csv(csv_buffer, index=False)
    csv_data = csv_buffer.getvalue()

    return csv_data, filename


def export_portfolio_excel(holdings_df, metrics, monthly_income, account_name="Portfolio", filename=None):
    """
    Export portfolio to multi-sheet Excel file with formatting.

    Args:
        holdings_df: DataFrame with holdings data
        metrics: Dictionary with portfolio metrics
        monthly_income: DataFrame with monthly income data
        account_name: Name of the account
        filename: Optional custom filename

    Returns:
        Tuple of (excel_bytes, filename) for download
    """
    if not OPENPYXL_AVAILABLE:
        return None, None

    if holdings_df is None or holdings_df.empty:
        return None, None

    # Auto-generate filename
    if filename is None:
        date_str = datetime.now().strftime("%Y-%m-%d")
        filename = f"portfolio_{account_name}_{date_str}.xlsx"

    # Create workbook
    wb = openpyxl.Workbook()

    # Define styles
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="1a365d", end_color="1a365d", fill_type="solid")
    border = Border(
        left=Side(style='thin', color='cccccc'),
        right=Side(style='thin', color='cccccc'),
        top=Side(style='thin', color='cccccc'),
        bottom=Side(style='thin', color='cccccc')
    )
    center_align = Alignment(horizontal='center', vertical='center')

    # === Sheet 1: Holdings ===
    ws1 = wb.active
    ws1.title = "Holdings"

    # Write holdings data
    export_cols = ['Ticker', 'Category', 'Shares', 'Price £', 'Value £', 'Weight %', 'P/L %', 'Trail Yield %', 'Est Annual £']
    available_cols = [c for c in export_cols if c in holdings_df.columns]
    export_df = holdings_df[available_cols].copy()

    for r_idx, row in enumerate(dataframe_to_rows(export_df, index=False, header=True), 1):
        for c_idx, value in enumerate(row, 1):
            cell = ws1.cell(row=r_idx, column=c_idx, value=value)
            cell.border = border
            if r_idx == 1:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = center_align

    # Auto-adjust column widths
    for col_idx, col in enumerate(available_cols, 1):
        max_len = max(len(str(col)), max(len(str(v)) for v in export_df[col].astype(str)))
        ws1.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = min(max_len + 2, 20)

    # === Sheet 2: Summary Metrics ===
    ws2 = wb.create_sheet("Summary")

    summary_data = [
        ["Portfolio Summary Report", ""],
        ["Generated", datetime.now().strftime("%Y-%m-%d %H:%M")],
        ["Account", account_name],
        ["", ""],
        ["Key Metrics", "Value"],
        ["Total Value", f"£{metrics.get('total_value', 0):,.0f}"],
        ["Total Cost", f"£{metrics.get('total_cost', 0):,.0f}"],
        ["Unrealised P/L", f"£{metrics.get('total_value', 0) - metrics.get('total_cost', 0):,.0f}"],
        ["P/L %", f"{((metrics.get('total_value', 0) - metrics.get('total_cost', 0)) / max(metrics.get('total_cost', 1), 1) * 100):+.1f}%"],
        ["", ""],
        ["Income Projections", "Value"],
        ["Projected Annual Income", f"£{metrics.get('projected_annual', 0):,.0f}"],
        ["Projected Monthly Income", f"£{metrics.get('projected_monthly', 0):,.0f}"],
        ["Yield on Value", f"{(metrics.get('projected_annual', 0) / max(metrics.get('total_value', 1), 1) * 100):.2f}%"],
        ["", ""],
        ["Risk Metrics", "Value"],
        ["Portfolio Volatility", f"{metrics.get('portfolio_vol', 0):.1f}%"],
        ["Average Beta", f"{metrics.get('portfolio_beta', 1):.2f}"],
        ["USD Exposure", f"{(metrics.get('usd_exposure', 0) / max(metrics.get('total_value', 1), 1) * 100):.1f}%"],
    ]

    for r_idx, row in enumerate(summary_data, 1):
        for c_idx, value in enumerate(row, 1):
            cell = ws2.cell(row=r_idx, column=c_idx, value=value)
            cell.border = border
            if r_idx == 1:
                cell.font = Font(bold=True, size=14)
            elif value in ["Key Metrics", "Income Projections", "Risk Metrics", "Value"]:
                cell.font = header_font
                cell.fill = header_fill

    ws2.column_dimensions['A'].width = 25
    ws2.column_dimensions['B'].width = 20

    # === Sheet 3: Monthly Income ===
    ws3 = wb.create_sheet("Monthly Income")

    if monthly_income is not None and not monthly_income.empty:
        for r_idx, row in enumerate(dataframe_to_rows(monthly_income, index=False, header=True), 1):
            for c_idx, value in enumerate(row, 1):
                cell = ws3.cell(row=r_idx, column=c_idx, value=value)
                cell.border = border
                if r_idx == 1:
                    cell.font = header_font
                    cell.fill = header_fill

        # Add summary stats at bottom
        stats_start = len(monthly_income) + 3
        col_name = 'Total Amount' if 'Total Amount' in monthly_income.columns else 'Total'
        ws3.cell(row=stats_start, column=1, value="Total YTD").font = Font(bold=True)
        ws3.cell(row=stats_start, column=2, value=f"£{monthly_income[col_name].sum():,.0f}")
        ws3.cell(row=stats_start + 1, column=1, value="Average Monthly").font = Font(bold=True)
        ws3.cell(row=stats_start + 1, column=2, value=f"£{monthly_income[col_name].mean():,.0f}")

    # Save to bytes
    excel_buffer = io.BytesIO()
    wb.save(excel_buffer)
    excel_bytes = excel_buffer.getvalue()

    return excel_bytes, filename


def generate_portfolio_report(holdings_df, metrics, monthly_income, account_name="Portfolio",
                              isa_metrics=None, sipp_metrics=None, include_charts=True):
    """
    Generate a professional HTML report for the portfolio.

    Args:
        holdings_df: DataFrame with holdings data
        metrics: Dictionary with portfolio metrics
        monthly_income: DataFrame with monthly income data
        account_name: Name of the account
        isa_metrics: Optional ISA-specific metrics
        sipp_metrics: Optional SIPP-specific metrics
        include_charts: Whether to include Plotly charts as embedded HTML

    Returns:
        Tuple of (html_content, filename) for download
    """
    if holdings_df is None or holdings_df.empty:
        return None, None

    date_str = datetime.now().strftime("%Y-%m-%d")
    time_str = datetime.now().strftime("%H:%M")
    filename = f"portfolio_report_{account_name}_{date_str}.html"

    # Calculate metrics
    total_value = metrics.get('total_value', 0)
    total_cost = metrics.get('total_cost', 0)
    pl_value = total_value - total_cost
    pl_pct = (pl_value / total_cost * 100) if total_cost > 0 else 0
    projected_annual = metrics.get('projected_annual', 0)
    yield_on_value = (projected_annual / total_value * 100) if total_value > 0 else 0

    # Category breakdown
    category_summary = holdings_df.groupby('Category')['Value £'].sum().reset_index()
    category_html = ""
    for _, row in category_summary.iterrows():
        pct = (row['Value £'] / total_value * 100) if total_value > 0 else 0
        category_html += f'''
        <div class="category-item">
            <span class="category-name">{row['Category'].title()}</span>
            <span class="category-value">£{row['Value £']:,.0f} ({pct:.1f}%)</span>
        </div>
        '''

    # Holdings table
    holdings_html = '''
    <table class="holdings-table">
        <thead>
            <tr>
                <th>Ticker</th>
                <th>Category</th>
                <th>Shares</th>
                <th>Price</th>
                <th>Value</th>
                <th>Weight</th>
                <th>P/L</th>
                <th>Yield</th>
                <th>Est. Annual</th>
            </tr>
        </thead>
        <tbody>
    '''

    for _, row in holdings_df.iterrows():
        pl_class = "positive" if row.get('P/L %', 0) >= 0 else "negative"
        holdings_html += f'''
        <tr>
            <td class="ticker">{row['Ticker']}</td>
            <td>{row['Category']}</td>
            <td>{row.get('Shares', 0):.2f}</td>
            <td>£{row.get('Price £', 0):.2f}</td>
            <td>£{row.get('Value £', 0):,}</td>
            <td>{row.get('Weight %', 0):.1f}%</td>
            <td class="{pl_class}">{row.get('P/L %', 0):+.1f}%</td>
            <td>{row.get('Trail Yield %', 0):.2f}%</td>
            <td>£{row.get('Est Annual £', 0):,}</td>
        </tr>
        '''

    holdings_html += '''
        </tbody>
    </table>
    '''

    # Monthly income table
    income_html = ""
    if monthly_income is not None and not monthly_income.empty:
        col_name = 'Total Amount' if 'Total Amount' in monthly_income.columns else 'Total'
        income_html = '''
        <table class="income-table">
            <thead>
                <tr>
                    <th>Month</th>
                    <th>Income</th>
                </tr>
            </thead>
            <tbody>
        '''
        for _, row in monthly_income.iterrows():
            income_html += f'''
            <tr>
                <td>{row['Month']}</td>
                <td>£{row[col_name]:,.2f}</td>
            </tr>
            '''
        income_html += '''
            </tbody>
        </table>
        '''

        ytd_income = monthly_income[col_name].sum()
        avg_monthly = monthly_income[col_name].mean()
        income_html += f'''
        <div class="income-summary">
            <div class="stat">
                <span class="label">Total YTD</span>
                <span class="value">£{ytd_income:,.0f}</span>
            </div>
            <div class="stat">
                <span class="label">Avg Monthly</span>
                <span class="value">£{avg_monthly:,.0f}</span>
            </div>
        </div>
        '''

    # Generate charts if requested and Plotly is available
    charts_html = ""
    if include_charts and PLOTLY_AVAILABLE:
        # Allocation pie chart
        fig_alloc = px.pie(holdings_df, values='Value £', names='Ticker',
                          title="Holdings Allocation", hole=0.4)
        fig_alloc.update_layout(height=400, margin=dict(t=50, b=50, l=50, r=50))
        alloc_chart = fig_alloc.to_html(full_html=False, include_plotlyjs='cdn')

        # Category pie chart
        fig_cat = px.pie(category_summary, values='Value £', names='Category',
                        title="Category Allocation", hole=0.4)
        fig_cat.update_layout(height=400, margin=dict(t=50, b=50, l=50, r=50))
        cat_chart = fig_cat.to_html(full_html=False, include_plotlyjs=False)

        charts_html = f'''
        <div class="charts-section">
            <h2>Portfolio Visualizations</h2>
            <div class="charts-grid">
                <div class="chart-container">
                    {alloc_chart}
                </div>
                <div class="chart-container">
                    {cat_chart}
                </div>
            </div>
        </div>
        '''

    # Build the complete HTML
    pl_class_str = 'positive' if pl_value >= 0 else 'negative'

    html_content = f'''<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>Portfolio Report - {account_name} - {date_str}</title>
    <style>
        :root {{
            --primary: #1a365d;
            --secondary: #2563eb;
            --success: #16a34a;
            --danger: #dc2626;
            --bg: #f8fafc;
            --surface: #ffffff;
            --text: #1e293b;
            --text-secondary: #64748b;
        }}
        * {{ margin: 0; padding: 0; box-sizing: border-box; }}
        body {{
            font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', Roboto, 'Helvetica Neue', Arial, sans-serif;
            background-color: var(--bg); color: var(--text); line-height: 1.6; padding: 2rem;
        }}
        .container {{ max-width: 1200px; margin: 0 auto; }}
        .header {{
            text-align: center; margin-bottom: 3rem; padding-bottom: 2rem;
            border-bottom: 3px solid var(--primary);
        }}
        .header h1 {{ font-size: 2.5rem; color: var(--primary); margin-bottom: 0.5rem; }}
        .header .subtitle {{ color: var(--text-secondary); font-size: 1.1rem; }}
        .header .date {{ color: var(--text-secondary); font-size: 0.9rem; margin-top: 0.5rem; }}
        .section {{
            background: var(--surface); border-radius: 12px; padding: 1.5rem;
            margin-bottom: 2rem; box-shadow: 0 4px 6px -1px rgba(0, 0, 0, 0.1);
        }}
        .section h2 {{
            color: var(--primary); font-size: 1.5rem; margin-bottom: 1.5rem;
            padding-bottom: 0.5rem; border-bottom: 2px solid var(--bg);
        }}
        .metrics-grid {{
            display: grid; grid-template-columns: repeat(auto-fit, minmax(200px, 1fr)); gap: 1.5rem;
        }}
        .metric-card {{
            background: linear-gradient(135deg, var(--primary), var(--secondary));
            color: white; padding: 1.5rem; border-radius: 10px; text-align: center;
        }}
        .metric-card .value {{ font-size: 2rem; font-weight: bold; display: block; }}
        .metric-card .label {{ font-size: 0.9rem; opacity: 0.9; }}
        .metric-card.positive {{ background: linear-gradient(135deg, var(--success), #22c55e); }}
        .metric-card.negative {{ background: linear-gradient(135deg, var(--danger), #f87171); }}
        .category-breakdown {{ display: flex; flex-direction: column; gap: 0.75rem; }}
        .category-item {{
            display: flex; justify-content: space-between; padding: 0.75rem 1rem;
            background: var(--bg); border-radius: 8px; border-left: 4px solid var(--secondary);
        }}
        .category-name {{ font-weight: 600; }}
        .category-value {{ color: var(--text-secondary); }}
        .holdings-table, .income-table {{ width: 100%; border-collapse: collapse; margin-top: 1rem; }}
        .holdings-table th, .holdings-table td, .income-table th, .income-table td {{
            padding: 0.75rem 1rem; text-align: left; border-bottom: 1px solid var(--bg);
        }}
        .holdings-table th, .income-table th {{
            background: var(--primary); color: white; font-weight: 600;
        }}
        .holdings-table tr:hover, .income-table tr:hover {{ background: var(--bg); }}
        .holdings-table .ticker {{ font-weight: 600; color: var(--primary); }}
        .positive {{ color: var(--success); font-weight: 600; }}
        .negative {{ color: var(--danger); font-weight: 600; }}
        .charts-section {{ margin-top: 2rem; }}
        .charts-grid {{
            display: grid; grid-template-columns: repeat(auto-fit, minmax(400px, 1fr)); gap: 2rem;
        }}
        .chart-container {{
            background: var(--surface); border-radius: 12px; padding: 1rem;
            box-shadow: 0 2px 4px rgba(0, 0, 0, 0.1);
        }}
        .income-summary {{
            display: flex; gap: 2rem; margin-top: 1.5rem; padding-top: 1rem;
            border-top: 2px solid var(--bg);
        }}
        .income-summary .stat {{ text-align: center; }}
        .income-summary .label {{ display: block; color: var(--text-secondary); font-size: 0.9rem; }}
        .income-summary .value {{ display: block; font-size: 1.5rem; font-weight: bold; color: var(--success); }}
        .footer {{
            text-align: center; padding: 2rem; color: var(--text-secondary);
            font-size: 0.9rem; border-top: 1px solid var(--bg); margin-top: 2rem;
        }}
        @media print {{
            body {{ padding: 1rem; }}
            .section {{ break-inside: avoid; }}
            .charts-section {{ break-before: page; }}
        }}
        @media (max-width: 768px) {{
            .metrics-grid {{ grid-template-columns: repeat(2, 1fr); }}
            .charts-grid {{ grid-template-columns: 1fr; }}
        }}
    </style>
</head>
<body>
    <div class="container">
        <div class="header">
            <h1>Portfolio Report</h1>
            <p class="subtitle">{account_name}</p>
            <p class="date">Generated on {date_str} at {time_str}</p>
        </div>

        <div class="section">
            <h2>Portfolio Summary</h2>
            <div class="metrics-grid">
                <div class="metric-card">
                    <span class="value">£{total_value:,.0f}</span>
                    <span class="label">Total Value</span>
                </div>
                <div class="metric-card {pl_class_str}">
                    <span class="value">{pl_pct:+.1f}%</span>
                    <span class="label">Total P/L (£{pl_value:,.0f})</span>
                </div>
                <div class="metric-card">
                    <span class="value">£{projected_annual:,.0f}</span>
                    <span class="label">Projected Annual Income</span>
                </div>
                <div class="metric-card">
                    <span class="value">{yield_on_value:.2f}%</span>
                    <span class="label">Yield on Value</span>
                </div>
            </div>
        </div>

        <div class="section">
            <h2>Category Allocation</h2>
            <div class="category-breakdown">
                {category_html}
            </div>
        </div>

        <div class="section">
            <h2>Holdings</h2>
            {holdings_html}
        </div>

        <div class="section">
            <h2>Monthly Dividend Income</h2>
            {income_html if income_html else '<p style="color: var(--text-secondary);">No dividend data available yet.</p>'}
        </div>

        {charts_html}

        <div class="footer">
            <p>This report was automatically generated by Freetrade Portfolio Dashboard</p>
            <p>Data is for informational purposes only and should not be considered financial advice.</p>
        </div>
    </div>
</body>
</html>'''

    return html_content, filename


def generate_pdf_report(holdings_df, metrics, monthly_income, account_name="Portfolio"):
    """
    Generate a PDF report (requires weasyprint).

    Returns:
        Tuple of (pdf_bytes, filename, error_message) - error_message is None on success
    """
    if not WEASYPRINT_AVAILABLE:
        return None, None, '''
PDF export requires the weasyprint library. To enable PDF export:

1. Install system dependencies (macOS):
   brew install cairo pango gdk-pixbuf libffi

2. Install system dependencies (Ubuntu/Debian):
   sudo apt-get install libpango-1.0-0 libpangocairo-1.0-0 libgdk-pixbuf2.0-0 libffi-dev shared-mime-info

3. Install Python package:
   pip install weasyprint

After installation, restart the dashboard to enable PDF export.
        '''

    # Generate HTML report first (without interactive charts)
    html_content, _ = generate_portfolio_report(
        holdings_df, metrics, monthly_income, account_name, include_charts=False
    )

    if html_content is None:
        return None, None, "No data available for PDF export."

    date_str = datetime.now().strftime("%Y-%m-%d")
    filename = f"portfolio_report_{account_name}_{date_str}.pdf"

    try:
        pdf_buffer = io.BytesIO()
        WeasyHTML(string=html_content).write_pdf(pdf_buffer)
        pdf_bytes = pdf_buffer.getvalue()
        return pdf_bytes, filename, None
    except Exception as e:
        return None, None, f"Error generating PDF: {str(e)}"


def render_export_panel(isa_metrics=None, sipp_metrics=None, isa_df=None, sipp_df=None,
                       get_monthly_dividends_func=None, get_combined_monthly_dividends_func=None):
    """
    Render a clean UI panel for export options.

    Args:
        isa_metrics: ISA portfolio metrics dictionary
        sipp_metrics: SIPP portfolio metrics dictionary
        isa_df: ISA transactions DataFrame
        sipp_df: SIPP transactions DataFrame
        get_monthly_dividends_func: Function to get monthly dividends from a DataFrame
        get_combined_monthly_dividends_func: Function to get combined monthly dividends
    """
    st.header("📤 Export & Reports")

    # Check if we have data
    has_isa = isa_metrics is not None
    has_sipp = sipp_metrics is not None

    if not has_isa and not has_sipp:
        st.warning("No portfolio data available to export.")
        return

    # Account selection
    account_options = []
    if has_isa and has_sipp:
        account_options = ["Combined (ISA + SIPP)", "ISA Only", "SIPP Only"]
    elif has_isa:
        account_options = ["ISA"]
    else:
        account_options = ["SIPP"]

    col1, col2 = st.columns([2, 1])

    with col1:
        selected_account = st.radio(
            "Select Account",
            account_options,
            horizontal=True,
            key="export_account"
        )

    with col2:
        export_format = st.radio(
            "Export Format",
            ["CSV", "Excel", "HTML Report", "PDF Report"],
            key="export_format"
        )

    # Determine which data to use
    if "Combined" in selected_account:
        # Combine ISA and SIPP data
        combined_df = pd.concat([
            isa_metrics['df'].assign(Account='ISA') if isa_metrics else pd.DataFrame(),
            sipp_metrics['df'].assign(Account='SIPP') if sipp_metrics else pd.DataFrame()
        ])
        combined_metrics = {
            'total_value': (isa_metrics.get('total_value', 0) if isa_metrics else 0) +
                          (sipp_metrics.get('total_value', 0) if sipp_metrics else 0),
            'total_cost': (isa_metrics.get('total_cost', 0) if isa_metrics else 0) +
                         (sipp_metrics.get('total_cost', 0) if sipp_metrics else 0),
            'projected_annual': (isa_metrics.get('projected_annual', 0) if isa_metrics else 0) +
                               (sipp_metrics.get('projected_annual', 0) if sipp_metrics else 0),
            'projected_monthly': (isa_metrics.get('projected_monthly', 0) if isa_metrics else 0) +
                                (sipp_metrics.get('projected_monthly', 0) if sipp_metrics else 0),
            'usd_exposure': (isa_metrics.get('usd_exposure', 0) if isa_metrics else 0) +
                           (sipp_metrics.get('usd_exposure', 0) if sipp_metrics else 0),
            'portfolio_vol': np.mean([
                isa_metrics.get('portfolio_vol', 0) if isa_metrics else 0,
                sipp_metrics.get('portfolio_vol', 0) if sipp_metrics else 0
            ]),
            'portfolio_beta': np.mean([
                isa_metrics.get('portfolio_beta', 1) if isa_metrics else 1,
                sipp_metrics.get('portfolio_beta', 1) if sipp_metrics else 1
            ]),
        }
        if get_combined_monthly_dividends_func:
            monthly_income = get_combined_monthly_dividends_func(isa_df, sipp_df)
        else:
            monthly_income = pd.DataFrame()
        account_name = "Combined"
        export_df = combined_df
        export_metrics = combined_metrics
    elif "ISA" in selected_account:
        export_df = isa_metrics['df'] if isa_metrics else pd.DataFrame()
        export_metrics = isa_metrics if isa_metrics else {}
        if get_monthly_dividends_func:
            monthly_income = get_monthly_dividends_func(isa_df)
        else:
            monthly_income = pd.DataFrame()
        account_name = "ISA"
    else:
        export_df = sipp_metrics['df'] if sipp_metrics else pd.DataFrame()
        export_metrics = sipp_metrics if sipp_metrics else {}
        if get_monthly_dividends_func:
            monthly_income = get_monthly_dividends_func(sipp_df)
        else:
            monthly_income = pd.DataFrame()
        account_name = "SIPP"

    st.divider()

    # Export options specific to format
    if export_format == "Excel" and not OPENPYXL_AVAILABLE:
        st.warning("Excel export requires openpyxl. Install with: `pip install openpyxl`")
        return

    if export_format == "PDF Report":
        if not WEASYPRINT_AVAILABLE:
            st.warning("PDF export requires weasyprint.")
            with st.expander("How to enable PDF export"):
                st.markdown("""
**macOS:**
```bash
brew install cairo pango gdk-pixbuf libffi
pip install weasyprint
```

**Ubuntu/Debian:**
```bash
sudo apt-get install libpango-1.0-0 libpangocairo-1.0-0 libgdk-pixbuf2.0-0 libffi-dev
pip install weasyprint
```
                """)
            return

    # Additional options for HTML/PDF reports
    include_charts = True
    if export_format in ["HTML Report", "PDF Report"]:
        include_charts = st.checkbox("Include interactive charts", value=True,
                                     disabled=(export_format == "PDF Report"))

    # Generate export button
    st.divider()

    if st.button("📥 Generate Export", type="primary", use_container_width=True):
        with st.spinner(f"Generating {export_format}..."):
            if export_format == "CSV":
                csv_data, filename = export_portfolio_csv(export_df, account_name)
                if csv_data:
                    st.success(f"CSV generated: {filename}")
                    st.download_button(
                        label="📥 Download CSV",
                        data=csv_data,
                        file_name=filename,
                        mime="text/csv",
                        use_container_width=True
                    )
                else:
                    st.error("Failed to generate CSV export.")

            elif export_format == "Excel":
                excel_bytes, filename = export_portfolio_excel(
                    export_df, export_metrics, monthly_income, account_name
                )
                if excel_bytes:
                    st.success(f"Excel file generated: {filename}")
                    st.download_button(
                        label="📥 Download Excel",
                        data=excel_bytes,
                        file_name=filename,
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        use_container_width=True
                    )
                else:
                    st.error("Failed to generate Excel export.")

            elif export_format == "HTML Report":
                html_content, filename = generate_portfolio_report(
                    export_df, export_metrics, monthly_income, account_name,
                    isa_metrics, sipp_metrics, include_charts
                )
                if html_content:
                    st.success(f"HTML report generated: {filename}")
                    st.download_button(
                        label="📥 Download HTML Report",
                        data=html_content,
                        file_name=filename,
                        mime="text/html",
                        use_container_width=True
                    )

                    # Preview option
                    with st.expander("Preview Report"):
                        st.components.v1.html(html_content, height=800, scrolling=True)
                else:
                    st.error("Failed to generate HTML report.")

            elif export_format == "PDF Report":
                pdf_bytes, filename, error = generate_pdf_report(
                    export_df, export_metrics, monthly_income, account_name
                )
                if pdf_bytes:
                    st.success(f"PDF report generated: {filename}")
                    st.download_button(
                        label="📥 Download PDF Report",
                        data=pdf_bytes,
                        file_name=filename,
                        mime="application/pdf",
                        use_container_width=True
                    )
                else:
                    st.error(error or "Failed to generate PDF report.")

    # Quick export buttons
    st.divider()
    st.subheader("Quick Export")

    col1, col2, col3, col4 = st.columns(4)

    with col1:
        csv_data, csv_filename = export_portfolio_csv(export_df, account_name)
        if csv_data:
            st.download_button(
                label="📄 CSV",
                data=csv_data,
                file_name=csv_filename,
                mime="text/csv",
                key="quick_csv"
            )

    with col2:
        if OPENPYXL_AVAILABLE:
            excel_bytes, excel_filename = export_portfolio_excel(
                export_df, export_metrics, monthly_income, account_name
            )
            if excel_bytes:
                st.download_button(
                    label="📊 Excel",
                    data=excel_bytes,
                    file_name=excel_filename,
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    key="quick_excel"
                )
        else:
            st.button("📊 Excel", disabled=True, help="Install openpyxl")

    with col3:
        html_content, html_filename = generate_portfolio_report(
            export_df, export_metrics, monthly_income, account_name, include_charts=True
        )
        if html_content:
            st.download_button(
                label="📰 HTML",
                data=html_content,
                file_name=html_filename,
                mime="text/html",
                key="quick_html"
            )

    with col4:
        if WEASYPRINT_AVAILABLE:
            pdf_bytes, pdf_filename, _ = generate_pdf_report(
                export_df, export_metrics, monthly_income, account_name
            )
            if pdf_bytes:
                st.download_button(
                    label="📕 PDF",
                    data=pdf_bytes,
                    file_name=pdf_filename,
                    mime="application/pdf",
                    key="quick_pdf"
                )
        else:
            st.button("📕 PDF", disabled=True, help="Install weasyprint")


def schedule_report_email(placeholder=True):
    """
    Placeholder UI for scheduling weekly/monthly email reports.

    Note: Full implementation requires SMTP configuration.
    """
    st.subheader("📬 Schedule Email Reports")

    st.info("""
    **Coming Soon: Automated Email Reports**

    Schedule weekly or monthly portfolio reports to be delivered directly to your inbox.
    """)

    with st.expander("Configure Email Reports (Preview)"):
        col1, col2 = st.columns(2)

        with col1:
            st.selectbox(
                "Report Frequency",
                ["Weekly (Monday)", "Weekly (Friday)", "Monthly (1st)", "Monthly (15th)"],
                key="email_frequency",
                disabled=True
            )

            st.text_input(
                "Email Address",
                placeholder="your@email.com",
                key="email_address",
                disabled=True
            )

        with col2:
            st.multiselect(
                "Include Accounts",
                ["ISA", "SIPP", "Combined"],
                default=["Combined"],
                key="email_accounts",
                disabled=True
            )

            st.selectbox(
                "Report Format",
                ["PDF", "HTML", "Excel"],
                key="email_format",
                disabled=True
            )

        st.divider()

        st.markdown("""
**To enable email reports, configure SMTP settings:**

```python
# Add to config.json:
{
    "smtp_server": "smtp.gmail.com",
    "smtp_port": 587,
    "smtp_user": "your@email.com",
    "smtp_password": "app-specific-password"
}
```

For Gmail, use an [App Password](https://support.google.com/accounts/answer/185833).
        """)

        st.button("💾 Save Email Settings", disabled=True, use_container_width=True)


def render_export_sidebar_section(isa_metrics=None, sipp_metrics=None, isa_df=None, sipp_df=None,
                                  get_monthly_dividends_func=None, get_combined_monthly_dividends_func=None):
    """
    Render a compact export section in the sidebar.
    """
    with st.sidebar:
        st.divider()
        st.subheader("📤 Quick Export")

        # Determine what data is available
        has_data = isa_metrics is not None or sipp_metrics is not None

        if not has_data:
            st.caption("Export available after data loads")
            return

        # Combine data for export
        if isa_metrics and sipp_metrics:
            combined_df = pd.concat([isa_metrics['df'], sipp_metrics['df']])
            combined_metrics = {
                'total_value': isa_metrics['total_value'] + sipp_metrics['total_value'],
                'total_cost': isa_metrics['total_cost'] + sipp_metrics['total_cost'],
                'projected_annual': isa_metrics['projected_annual'] + sipp_metrics['projected_annual'],
                'projected_monthly': isa_metrics['projected_monthly'] + sipp_metrics['projected_monthly'],
                'usd_exposure': isa_metrics['usd_exposure'] + sipp_metrics['usd_exposure'],
                'portfolio_vol': (isa_metrics['portfolio_vol'] + sipp_metrics['portfolio_vol']) / 2,
                'portfolio_beta': (isa_metrics['portfolio_beta'] + sipp_metrics['portfolio_beta']) / 2,
            }
            if get_combined_monthly_dividends_func:
                monthly_income = get_combined_monthly_dividends_func(isa_df, sipp_df)
            else:
                monthly_income = pd.DataFrame()
            account_name = "Combined"
        elif isa_metrics:
            combined_df = isa_metrics['df']
            combined_metrics = isa_metrics
            if get_monthly_dividends_func:
                monthly_income = get_monthly_dividends_func(isa_df)
            else:
                monthly_income = pd.DataFrame()
            account_name = "ISA"
        else:
            combined_df = sipp_metrics['df']
            combined_metrics = sipp_metrics
            if get_monthly_dividends_func:
                monthly_income = get_monthly_dividends_func(sipp_df)
            else:
                monthly_income = pd.DataFrame()
            account_name = "SIPP"

        # Quick export buttons
        col1, col2 = st.columns(2)

        with col1:
            csv_data, csv_filename = export_portfolio_csv(combined_df, account_name)
            if csv_data:
                st.download_button(
                    "📄 CSV",
                    data=csv_data,
                    file_name=csv_filename,
                    mime="text/csv",
                    use_container_width=True,
                    key="sidebar_csv"
                )

        with col2:
            html_content, html_filename = generate_portfolio_report(
                combined_df, combined_metrics, monthly_income, account_name, include_charts=False
            )
            if html_content:
                st.download_button(
                    "📰 Report",
                    data=html_content,
                    file_name=html_filename,
                    mime="text/html",
                    use_container_width=True,
                    key="sidebar_html"
                )
